# -*- coding: utf-8 -*-
import os
from datetime import datetime

import pythoncom
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from win32com.client import Dispatch

from ap_service.settings import REPORT_DIR, TEMPLATE_DIR
from utils.date_utils import parse_date_from_string, convert_datetime_to_string


def create_xlsx_file_using_template(report_name, number_sheet, template='template.xlsx'):
    """
    :param number_sheet:
    :param export_type: 2 create a file report; 3 create all reports
    :param list_date:
    :param templates: None
    :return: list reports
    """
    # create file xlsx
    wb = Workbook()
    path_report = os.path.join(REPORT_DIR, report_name)
    wb.save(path_report)

    path_template = os.path.join(TEMPLATE_DIR, template)

    # Copy templates to files report
    pythoncom.CoInitialize()
    xl = Dispatch('Excel.Application')
    # You can remove this line if you don't want the Excel application to be visible
    xl.Visible = True

    wb1 = xl.Workbooks.Open(Filename=path_template)
    wb2 = xl.Workbooks.Open(Filename=path_report)

    ws1 = wb1.Worksheets(1)
    ws1.Copy(Before=wb2.Worksheets(1))

    wb2.Close(SaveChanges=True)
    xl.Quit()

    wb = load_workbook(path_report)
    if number_sheet > 1:
        # Load report
        number = 2

        while number_sheet >= number:
            # Get Sheet
            source = wb.get_sheet_by_name('01')
            # Copy sheet
            target = wb.copy_worksheet(source)

            # Rename sheet copy
            name_sheet = wb.get_sheet_by_name('01 Copy')
            name_new_sheet = "%02d" % (number,)
            name_sheet.title = name_new_sheet
            wb.save(path_report)
            number += 1

    wb.remove(wb.get_sheet_by_name('Sheet'))
    wb.save(path_report)

    return path_report

# number sheet

def update_xlsx_file(report_data, path_report, index_sheet):
    # Create a workbook and add a worksheet.

    workbook = load_workbook(path_report)
    data = report_data
    # Open sheet index in file
    sheet = workbook.worksheets[index_sheet]

    ft = Font(name='Times New Roman',
              size=13,
              italic=True,
              strike=False,
              )
    day, month, year = "", "", ""

    if isinstance(data['created_at'], datetime) :
        date = convert_datetime_to_string(data['created_at'], type_format=2)
        date = parse_date_from_string(date)
        year = date.strftime("%Y")
        month = date.strftime("%m")
        day = date.strftime("%d")
    elif isinstance(data['created_at'],str):
        date = parse_date_from_string(data['created_at'])
        year = date.strftime("%Y")
        month = date.strftime("%m")
        day = date.strftime("%d")

    sheet['A2'] = "Ngày {} tháng {} năm {}".format(day, month, year)
    a2 = sheet['A2']
    a2.font = ft

    # A Phòng HÀNH CHÍNH - QUẢN TRỊ
    # I. Bộ phận hướng dẫn + tổng đài hỗ trợ 1900558826
    sheet['D7'] = data['a_i_1_amount']
    sheet['E7'] = data['a_i_1_note']
    sheet['D11'] = data['a_i_2_1_amount']
    sheet['E11'] = data['a_i_2_1_note']
    sheet['D12'] = data['a_i_2_2_amount']
    sheet['E12'] = data['a_i_2_2_note']
    sheet['D10'] = sheet['D11'].value + sheet['D12'].value
    sheet['E10'] = data['a_i_2_note']
    sheet['D14'] = data['a_i_2_3_amount']
    sheet['E14'] = data['a_i_2_3_note']
    sheet['D15'] = data['a_i_2_4_amount']
    sheet['E15'] = data['a_i_2_4_note']
    sheet['D13'] = sheet['D14'].value + sheet['D15'].value
    sheet['D8'] = sheet['D10'].value + sheet['D13'].value
    sheet['D16'] = data['a_i_3_amount']
    sheet['E16'] = data['a_i_3_note']
    sheet['D17'] = data['a_i_4_amount']
    sheet['E17'] = data['a_i_4_note']
    # II. Bộ phận thu phí, lệ phí
    sheet['D19'] = data['a_ii_1_amount']
    sheet['E19'] = data['a_ii_1_note']
    sheet['D22'] = data['a_ii_2_1_amount']
    sheet['E22'] = data['a_ii_2_1_note']
    sheet['D23'] = data['a_ii_2_2_amount']
    sheet['E23'] = data['a_ii_2_2_note']
    sheet['D24'] = data['a_ii_2_3_amount']
    sheet['E24'] = data['a_ii_2_3_note']
    sheet['D25'] = data['a_ii_2_4_amount']
    sheet['E25'] = data['a_ii_2_4_note']
    sheet['D26'] = data['a_ii_2_5_amount']
    sheet['E26'] = data['a_ii_2_5_note']
    sheet['D20'] = sheet['D22'].value + sheet['D23'].value + sheet['D24'].value + sheet['D25'].value + sheet[
        'D26'].value
    sheet['D27'] = data['a_ii_3_amount']
    sheet['E27'] = data['a_ii_3_note']
    # III. Hoạt động của các bộ phận dịch vụ hỗ trợ
    sheet['D29'] = data['a_iii_1_1_amount']
    sheet['E29'] = data['a_iii_1_1_note']
    sheet['D30'] = data['a_iii_1_2_amount']
    sheet['E30'] = data['a_iii_1_2_note']
    sheet['D33'] = data['a_iii_2_1_amount']
    sheet['E33'] = data['a_iii_2_1_note']
    sheet['D34'] = data['a_iii_2_2_mount']
    sheet['E34'] = data['a_iii_2_2_note']
    sheet['D35'] = data['a_iii_2_3_mount']
    sheet['E35'] = data['a_iii_2_3_note']
    sheet['D31'] = sheet['D33'].value + sheet['D34'].value + sheet['D35'].value
    sheet['E31'] = data['a_iii_2_note']
    sheet['D38'] = data['a_iii_3_1_mount']
    sheet['E38'] = data['a_iii_3_1_note']
    sheet['D39'] = data['a_iii_3_2_mount']
    sheet['E39'] = data['a_iii_3_2_note']
    sheet['D40'] = data['a_iii_3_3_mount']
    sheet['E40'] = data['a_iii_3_3_note']
    sheet['D36'] = sheet['D38'].value + sheet['D39'].value + sheet['D40'].value
    sheet['E36'] = data['a_iii_3_note']
    sheet['D41'] = data['a_iii_4_mount']
    sheet['E41'] = data['a_iii_4_note']
    sheet['D42'] = data['a_iii_5_mount']
    sheet['E42'] = data['a_iii_5_note']
    sheet['D43'] = data['a_iii_6_mount']
    sheet['E43'] = data['a_iii_6_note']
    sheet['D44'] = data['a_iii_7_1_mount']
    sheet['E44'] = data['a_iii_7_1_note']
    sheet['D45'] = data['a_iii_7_2_mount']
    sheet['E45'] = data['a_iii_7_2_note']
    # B. PHÒNG TIẾP NHẬN VÀ GIẢI QUYẾT
    # I. Tiếp nhận, giải quyết, trả kết quả TTHC tại Trung tâm Phục vụ hành chính công tỉnh
    sheet['D50'] = data['b_i_1_1_amount']
    sheet['E50'] = data['b_i_1_1_note']
    sheet['D51'] = data['b_i_1_2_amount']
    sheet['E51'] = data['b_i_1_2_note']
    sheet['D52'] = data['b_i_1_3_amount']
    sheet['E52'] = data['b_i_1_3_note']
    sheet['D48'] = sheet['D50'].value + sheet['D51'].value + sheet['D52'].value
    sheet['E48'] = data['b_i_1_note']
    sheet['D55'] = data['b_i_2_1_amount']
    sheet['E55'] = data['b_i_2_1_note']
    sheet['D56'] = data['b_i_2_2_amount']
    sheet['E56'] = data['b_i_2_2_note']
    sheet['D57'] = data['b_i_2_3_amount']
    sheet['E57'] = data['b_i_2_3_note']
    sheet['D53'] = sheet['D55'].value + sheet['D56'].value + sheet['D57'].value
    sheet['E53'] = data['b_i_2_note']
    sheet['D60'] = data['b_i_3_1_amount']
    sheet['E60'] = data['b_i_3_1_note']
    sheet['D61'] = data['b_i_3_2_amount']
    sheet['E61'] = data['b_i_3_2_note']
    sheet['D62'] = data['b_i_3_3_amount']
    sheet['E62'] = data['b_i_3_3_note']
    sheet['D58'] = sheet['D60'].value + sheet['D61'].value + sheet['D62'].value
    sheet['E58'] = data['b_i_3_note']
    # .II Các nội dung khác
    sheet['D64'] = data['b_ii_1_1_amount']
    sheet['E64'] = data['b_ii_1_1_note']
    sheet['D65'] = data['b_ii_1_2_amount']
    sheet['E65'] = data['b_ii_1_2_note']
    sheet['D66'] = data['b_ii_2_1_amount']
    sheet['E66'] = data['b_ii_2_1_note']
    sheet['D67'] = data['b_ii_2_2_amount']
    sheet['E67'] = data['b_ii_2_2_note']
    sheet['D68'] = data['b_ii_3_amount']
    sheet['E68'] = data['b_ii_3_note']
    sheet['D69'] = data['b_ii_4_amount']
    sheet['E69'] = data['b_ii_4_note']
    # C. PHÒNG KẾ HOẠCH TỔNG HỢP
    # I. Tiếp nhận, giải quyết, trả kết quả TTHC tại Trung tâm Phục vụ hành chính công tỉnh
    sheet['D72'] = data['c_i_1_1_amount']
    sheet['E72'] = data['c_i_1_1_note']
    sheet['D73'] = data['c_i_1_2_amount']
    sheet['E73'] = data['c_i_1_2_note']
    sheet['D77'] = data['c_i_2_1_amount']
    sheet['E77'] = data['c_i_2_1_note']
    sheet['D78'] = data['c_i_2_2_amount']
    sheet['E78'] = data['c_i_2_2_note']
    sheet['D79'] = data['c_i_2_3_amount']
    sheet['E79'] = data['c_i_2_3_note']
    sheet['D75'] = sheet['D77'].value + sheet['D78'].value + sheet['D79'].value
    sheet['E75'] = data['c_i_2_note']
    sheet['D80'] = data['c_i_3_1_amount']
    sheet['E80'] = data['c_i_3_1_note']
    sheet['D81'] = data['c_i_3_2_amount']
    sheet['E81'] = data['c_i_3_2_note']
    sheet['D82'] = data['c_i_4_1_amount']
    sheet['E82'] = data['c_i_4_1_note']
    sheet['D83'] = data['c_i_4_2_amount']
    sheet['E83'] = data['c_i_4_2_note']
    # II. Tiếp nhận, giải quyết TTHC tại Bộ phận tiếp nhận và trả kết quả cấp xã
    sheet['D85'] = data['c_ii_1_amount']
    sheet['e85'] = data['c_ii_1_note']
    sheet['D89'] = data['c_ii_2_1_amount']
    sheet['E89'] = data['c_ii_2_1_note']
    sheet['D90'] = data['c_ii_2_2_amount']
    sheet['E90'] = data['c_ii_2_2_note']
    sheet['D91'] = data['c_ii_2_3_amount']
    sheet['E91'] = data['c_ii_2_3_note']
    sheet['D87'] = sheet['D89'].value + sheet['D90'].value + sheet['D91'].value
    sheet['E87'] = data['c_ii_2_note']
    # III. Các nội dung khác
    sheet['D93'] = data['c_iii_1_amount']
    sheet['E93'] = data['c_iii_1_note']
    sheet['D94'] = data['c_iii_2_amount']
    sheet['E94'] = data['c_iii_2_note']
    sheet['D95'] = data['c_iii_3_amount']
    sheet['E95'] = data['c_iii_3_note']
    sheet['D96'] = data['c_iii_4_amount']
    sheet['E96'] = data['c_iii_4_note']
    sheet['D97'] = data['c_iii_5_amount']
    sheet['E97'] = data['c_iii_5_note']
    # D. PHÒNG KIỂM TRA GIÁM SAT
    # I. Tiếp nhận, trả kết quả giải quyết TTHC tại các Trung tâm PVHCC cấp huyện
    sheet['D100'] = data['d_i_1_amount']
    # sheet['E100'] = data['d_i_1_note']
    sheet['D101'] = data['d_i_2_amount']
    sheet['E101'] = data['d_i_2_note']
    sheet['D102'] = data['d_i_3_amount']
    sheet['E102'] = data['d_i_3_note']
    sheet['D103'] = data['d_i_4_amount']
    sheet['E103'] = data['d_i_4_note']
    sheet['D104'] = data['d_i_5_amount']
    sheet['E104'] = data['d_i_5_note']
    sheet['D105'] = data['d_i_6_amount']
    sheet['E105'] = data['d_i_6_note']
    # II. Khảo sát, đánh giá sự hài lòng của tổ chức, công dân
    sheet['D109'] = data['d_ii_1_1_amount']
    sheet['E109'] = data['d_ii_1_1_note']
    sheet['D110'] = data['d_ii_1_2_amount']
    sheet['E110'] = data['d_ii_1_2_note']
    sheet['D111'] = data['d_ii_1_3_amount']
    sheet['E111'] = data['d_ii_1_3_note']
    sheet['D107'] = sheet['D109'].value + sheet['D110'].value + sheet['D111'].value
    sheet['E107'] = data['d_ii_1_note']
    sheet['D113'] = data['d_ii_1_4_amount']
    sheet['E113'] = data['d_ii_1_4_note']
    if sheet['D107'].value > 0:
        sheet['D114'] = round(sheet['D113'].value * 100 / float(sheet['D107'].value), 2)
        sheet['D114'] = sheet['D107'].value
        sheet['D115'] = data['d_ii_1_5_amount']
        sheet['E115'] = data['d_ii_1_5_note']
        sheet['D116'] = round(sheet['D115'].value * 100 / float(sheet['D107'].value), 2)
        sheet['D117'] = data['d_ii_1_6_amount']
        sheet['E117'] = data['d_ii_1_6_note']
        sheet['D118'] = round(sheet['D117'].value * 100 / float(sheet['D107'].value), 2)
        sheet['D119'] = data['d_ii_1_7_amount']
        sheet['E119'] = data['d_ii_1_7_note']
        sheet['D120'] = round(100 - (sheet['D114'].value + sheet['D116'].value + sheet['D118'].value), 2)
    # III. Tiếp nhận, xử lý các phản ánh, kiến nghị, khiếu nại, tố cáo của tổ chức, công dân
    sheet['D124'] = data['d_iii_1_1_amount']
    sheet['E124'] = data['d_iii_1_1_note']
    sheet['D125'] = data['d_iii_1_2_amount']
    sheet['E125'] = data['d_iii_1_2_note']
    sheet['D126'] = data['d_iii_1_3_amount']
    sheet['E126'] = data['d_iii_1_3_note']
    sheet['D127'] = data['d_iii_1_4_amount']
    sheet['E127'] = data['d_iii_1_4_note']
    sheet['D123'] = sheet['D124'].value + sheet['D125'].value + sheet['D126'].value + sheet['D127'].value
    sheet['E123'] = data['d_iii_1_a_note']
    sheet['D129'] = data['d_iii_1_5_amount']
    sheet['E129'] = data['d_iii_1_5_note']
    sheet['D130'] = data['d_iii_1_6_amount']
    sheet['E130'] = data['d_iii_1_6_note']
    sheet['D131'] = data['d_iii_1_7_amount']
    sheet['E131'] = data['d_iii_1_7_note']
    sheet['D132'] = data['d_iii_1_8_amount']
    sheet['E132'] = data['d_iii_1_8_note']
    sheet['D128'] = sheet['D129'].value + sheet['D130'].value + sheet['D131'].value + sheet['D132'].value
    sheet['E128'] = data['d_iii_1_b_note']
    sheet['D135'] = data['d_iii_2_1_amount']
    sheet['E135'] = data['d_iii_2_1_note']
    sheet['D136'] = data['d_iii_2_2_amount']
    sheet['E136'] = data['d_iii_2_2_note']
    sheet['D137'] = data['d_iii_2_3_amount']
    sheet['E137'] = data['d_iii_2_3_note']
    sheet['D134'] = sheet['D135'].value + sheet['D136'].value + sheet['D137'].value
    sheet['E134'] = data['d_iii_2_a_note']
    sheet['D139'] = data['d_iii_2_4_amount']
    sheet['E139'] = data['d_iii_2_4_note']
    sheet['D140'] = data['d_iii_2_5_amount']
    sheet['E140'] = data['d_iii_2_5_note']
    sheet['D141'] = data['d_iii_2_6_amount']
    sheet['E141'] = data['d_iii_2_6_note']
    sheet['D138'] = sheet['D139'].value + sheet['D140'].value + sheet['D141'].value
    sheet['E138'] = data['d_iii_2_b_note']
    workbook.save(path_report)
    return path_report


def generate_report_name(export_type, start_date=None, end_date=None, created_at=None):
    if export_type == 0:
        start_date_str = start_date.strftime("%Y_%m_%d")
        end_date_str = end_date.strftime("%Y_%m_%d")
        report_name = "Report_summarized_from_{}_to_{}.xlsx".format(start_date_str, end_date_str)
        return report_name
    elif export_type == 1:
        start_date_str = start_date.strftime("%Y_%m_%d")
        end_date_str = end_date.strftime("%Y_%m_%d")
        report_name = "Report_from_{}_to_{}.xlsx".format(start_date_str, end_date_str)
        return report_name
    elif export_type == 2:
        date_str = created_at.strftime("%Y_%m_%d")
        report_name = "Report_{}.xlsx".format(date_str)
        return report_name

def is_file_existed(file_name, dir_filename=REPORT_DIR):
    if file_name:
        path_report = os.path.join(dir_filename, file_name)
        is_file = os.path.isfile(path_report)
        if is_file:
            return True

    return False


def rename_file_existing(file_name, new_name, dir_filename=REPORT_DIR):
    if is_file_existed(file_name, dir_filename):
        if len(new_name) > 0:
            old_path_report = os.path.join(REPORT_DIR, file_name)
            new_path_report = os.path.join(REPORT_DIR, new_name)
            os.rename(old_path_report, new_path_report)
            return new_name

def create_new_name_for_xlsx_file(file_name, file_name_extension):
    if len(file_name) > 0 and len(file_name_extension) > 0:
        new_file_name = "{}_old_{}.xlsx".format(file_name[:-5], file_name_extension)

        return new_file_name
